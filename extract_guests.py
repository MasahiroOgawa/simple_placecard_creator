#!/usr/bin/env python3
"""Extract all guest names from ゲスト一覧.xlsx including 連名 and output CSV."""

import argparse
import csv
import re

import openpyxl


def format_camelcase(raw):
    """McDermottEthan -> McDermott Ethan, BennasarJorge -> Bennasar Jorge."""
    return re.sub(r'(?<![Mm]c)(?<![Mm]ac)(?<=[a-z])([A-Z])', r' \1', raw)


def extract_guests(xlsx_path, special_cases=None):
    """Extract all attending guests + 連名 from the xlsx file.

    Args:
        xlsx_path: Path to the ゲスト一覧.xlsx file.
        special_cases: Optional dict of {guest_id: handler_function} for data quirks.

    Returns:
        List of dicts with 'display' and 'furigana' keys.
    """
    special_cases = special_cases or {}
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['ゲスト情報']

    all_participants = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        attendance = row[1].value  # B: 出欠情報
        if attendance != 'ご出席':
            continue

        guest_name = str(row[2].value or '').strip()
        furigana = str(row[3].value or '').strip()
        guest_id = row[0].value

        # Collect 連名 pairs (name, furigana)
        renmei_pairs = []
        for col_idx in [12, 14, 16, 18]:  # M, O, Q, S (0-indexed)
            name = row[col_idx].value
            furi = row[col_idx + 1].value if col_idx + 1 < len(row) else None
            if name:
                renmei_pairs.append((str(name).strip(), str(furi or '').strip()))

        # Apply special case handler if registered
        if guest_id in special_cases:
            results = special_cases[guest_id](guest_name, furigana, renmei_pairs)
            all_participants.extend(results)
            continue

        # Main guest
        if re.match(r'^[A-Za-z]+$', guest_name):
            display = format_camelcase(guest_name)
        else:
            display = guest_name
        all_participants.append({'display': display, 'furigana': furigana})

        # 連名
        for name, furi in renmei_pairs:
            if '/' in name:
                parts = name.split('/')
                display = f'{parts[1]} {parts[0]}'
            elif re.match(r'^[A-Za-z]+$', name):
                display = format_camelcase(name)
            else:
                display = name
            all_participants.append({'display': display, 'furigana': furi})

    return all_participants


def save_csv(participants, csv_path):
    """Save participant list as CSV."""
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['No', '表示名', '様付き', 'ふりがな'])
        for i, p in enumerate(participants):
            writer.writerow([i + 1, p['display'], f'{p["display"]}様', p['furigana']])


# === Special case handlers for this specific guest list ===

def handle_kouno_family(guest_name, furigana, renmei_pairs):
    """河野 family: main entry has typo, spouse in furigana field, 連名 are first-name only."""
    results = [
        {'display': '河野一弘', 'furigana': 'こうのかずひろ'},
        {'display': '河野クリステル', 'furigana': 'こうのくりすてる'},
    ]
    for name, furi in renmei_pairs:
        results.append({'display': f'河野{name}', 'furigana': f'こうの{furi}'})
    return results


def handle_ogawa_family(guest_name, furigana, renmei_pairs):
    """小川敏弘 family: 連名 field has family name, furigana has given name."""
    results = [{'display': guest_name, 'furigana': furigana}]
    for name, furi in renmei_pairs:
        results.append({'display': f'小川{furi}', 'furigana': f'{name}{furi}'})
    return results


SPECIAL_CASES = {
    2720955: handle_kouno_family,
    2675739: handle_ogawa_family,
}


def main():
    parser = argparse.ArgumentParser(description='Extract guest names from ゲスト一覧.xlsx')
    parser.add_argument('xlsx', help='Path to ゲスト一覧.xlsx')
    parser.add_argument('-o', '--output', default=None, help='Output CSV path (default: same dir as xlsx)')
    args = parser.parse_args()

    participants = extract_guests(args.xlsx, special_cases=SPECIAL_CASES)

    csv_path = args.output or args.xlsx.rsplit('.', 1)[0] + '_全参加者.csv'
    save_csv(participants, csv_path)

    print(f'Total participants: {len(participants)}')
    for i, p in enumerate(participants):
        print(f'{i+1:2d}. {p["display"]}様')
    print(f'\nCSV saved: {csv_path}')

    return participants


if __name__ == '__main__':
    main()
